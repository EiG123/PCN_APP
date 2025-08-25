import os
import logging
from collections import defaultdict
from tqdm import tqdm

import pandas as pd
import xml.etree.ElementTree as ET

from shapely.geometry import LineString, Point, MultiLineString, GeometryCollection
from shapely.ops import unary_union, transform

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pyproj import CRS, Transformer
from datetime import datetime

def write_results_to_excel(points_df, redline_summary, threshold_m, output_path=None, use_detail_count=False):
    """
    เขียนผลไปเป็น Excel:
      - sheet 'points_summary' = สรุปเส้น + นับ Close Action, Confirm, Revise พร้อม hyperlink
      - sheet per redline = รายละเอียด (ticket, sign, action/อื่นๆ)
      - ชื่อไฟล์จะใส่ threshold_m และวันที่เวลาปัจจุบัน
    
    Args:
        use_detail_count (bool): ถ้า True ใช้ points_by_details, ถ้า False ใช้ points (coordinate-based)
    """
    # ตั้งชื่อไฟล์ถ้าไม่ได้ส่งมา
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        count_type = "details" if use_detail_count else "coords"
        output_path = f"results_points_redlines_{threshold_m}m_{count_type}_{timestamp}.xlsx"

    # สร้าง DataFrame สรุป
    summary_rows = []
    for rl_name, info in redline_summary.items():
        # เลือกใช้ points แบบไหน สำหรับการนับใน summary
        # if use_detail_count:
        pts_for_summary = info['raw_matches']
        # else:
        #     pts_for_summary = info['points']
        ##############################################################################
        
        # ใช้ raw_matches สำหรับการแสดงจำนวนจริงใน detail sheet
        raw_matches = info['raw_matches']
        
        if not pts_for_summary:
            summary_rows.append({
                "เส้นสายไฟ": rl_name,
                "จำนวนจุดไม่ซ้ำ": 0,
                "จำนวนจุดรวมซ้ำ": 0,
                "Close Action": 0,
                "Confirm": 0,
                "Revise": 0,
                "อื่นๆ": 0,
                "ระยะเฉลี่ย (m)": 0
            })
            continue
        
        # ใช้ pts_for_summary สำหรับการนับประเภท (เพื่อไม่ให้นับซ้ำ)
        df_summary = pd.DataFrame(pts_for_summary)
        df_raw = pd.DataFrame(raw_matches) if raw_matches else pd.DataFrame()
        
        # นับแต่ละประเภทจาก summary data (ไม่ซ้ำ)
        close_action_count = (df_summary['sign'] == 'Close Action').sum() if 'sign' in df_summary.columns else 0
        confirm_count = (df_summary['sign'] == 'Confirm').sum() if 'sign' in df_summary.columns else 0
        revise_count = (df_summary['sign'] == 'Revise').sum() if 'sign' in df_summary.columns else 0
        
        # จุดที่เหลือ (ไม่ใช่ 3 ประเภทหลัก)
        main_categories = ['Close Action', 'Confirm', 'Revise']
        other_count = len(df_summary) - close_action_count - confirm_count - revise_count
        
        # จุดที่เหลือ (ไม่ใช่ 3 ประเภทหลัก)
        other_count = len(df_summary) - close_action_count - confirm_count - revise_count

        # คำนวณระยะเฉลี่ย
        avg_distance = df_raw['distance_m'].mean() if 'distance_m' in df_raw.columns and len(df_raw) > 0 else 0

        summary_rows.append({
            "เส้นสายไฟ": rl_name,
            "จำนวนจุดทั้งหมด": len(df_summary),
            "Close Action": close_action_count,
            "Confirm": confirm_count,
            "Revise": revise_count,
            "อื่นๆ": other_count,
            "ระยะเฉลี่ย (m)": round(avg_distance, 2)
        })

    summary_df = pd.DataFrame(summary_rows)
    
    total_row = {
        "เส้นสายไฟ": "รวมทั้งหมด",
        "จำนวนจุดทั้งหมด": summary_df["จำนวนจุดทั้งหมด"].sum(),
        "Close Action": summary_df["Close Action"].sum(),
        "Confirm": summary_df["Confirm"].sum(),
        "Revise": summary_df["Revise"].sum(),
        "อื่นๆ": summary_df["อื่นๆ"].sum(),
        "ระยะเฉลี่ย (m)": round(summary_df["ระยะเฉลี่ย (m)"].mean(), 2)
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

    # เขียนลง Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # เขียน Summary
        summary_df.to_excel(writer, sheet_name="points_summary", index=False)

        # เขียนแต่ละเส้น - ใช้ raw_matches (ข้อมูลทั้งหมดรวมซ้ำ)
        for rl_name, info in redline_summary.items():
            # ใช้ raw_matches ที่มีข้อมูลทั้งหมดรวมซ้ำ
            all_matches = info['raw_matches']
            if not all_matches:
                continue

            df = pd.DataFrame(all_matches)
            
            # เรียงข้อมูลตาม distance_m
            if 'distance_m' in df.columns:
                df = df.sort_values('distance_m')

            # sheet name จำกัด 31 chars และไม่ใช้อักขระพิเศษ
            safe_name = rl_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            sheet_name = (safe_name[:28] + '...') if len(safe_name) > 31 else safe_name

            # ตรวจชื่อซ้ำ
            existing_sheets = writer.book.sheetnames
            if sheet_name in existing_sheets:
                suffix = 1
                base_name = safe_name[:25] if len(safe_name) > 25 else safe_name
                while sheet_name in existing_sheets:
                    sheet_name = f"{base_name}_{suffix}"
                    suffix += 1

            # เขียนรายละเอียด
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # เขียนข้อมูลสถิติเพิ่มเติม
        stats_data = []
        for rl_name, info in redline_summary.items():
            stats_data.append({
                "เส้นสายไฟ": rl_name,
                "Count by Coordinates": info['count_by_coords'],
                "Count by Details": info['count_by_details'], 
                "Total Matches": info['total_matches'],
                "Duplicate Rate (%)": round(
                    ((info['total_matches'] - info['count_by_coords']) / info['total_matches'] * 100) 
                    if info['total_matches'] > 0 else 0, 2
                )
            })
        
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name="statistics", index=False)

    # โหลด workbook เพื่อเพิ่ม Hyperlink และจัดรูปแบบ
    wb = load_workbook(output_path)
    ws_summary = wb["points_summary"]

    # เพิ่ม Hyperlink (ยกเว้นแถวสุดท้ายที่เป็น "รวมทั้งหมด")
    for row_idx in range(2, len(summary_df)):  # ไม่รวมแถวสุดท้าย
        rl_name = ws_summary.cell(row=row_idx, column=1).value
        if rl_name == "รวมทั้งหมด":
            continue
            
        # ค้นหา sheet name ที่ตรงกับ rl_name หรือแบบตัด
        target_sheet = None
        safe_name = rl_name.replace('/', '_').replace('\\', '_').replace(':', '_')
        
        if safe_name in wb.sheetnames:
            target_sheet = safe_name
        else:
            # ค้นหาชื่อที่ตัดแล้ว
            for sname in wb.sheetnames:
                if sname.startswith(safe_name[:25]):
                    target_sheet = sname
                    break

        if target_sheet:
            cell = ws_summary.cell(row=row_idx, column=1)
            cell.hyperlink = f"#'{target_sheet}'!A1"
            cell.style = "Hyperlink"

    # # จัดรูปแบบแถว "รวมทั้งหมด"
    # total_row_idx = len(summary_df) + 1
    # for col in range(1, ws_summary.max_column + 1):
    #     cell = ws_summary.cell(row=total_row_idx, column=col)
    #     cell.font = Font(bold=True)
    #     cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    # # ปรับความกว้างคอลัมน์
    # for col in range(1, ws_summary.max_column + 1):
    #     max_len = 0
    #     col_letter = get_column_letter(col)
    #     for cell in ws_summary[col_letter]:
    #         try:
    #             if cell.value:
    #                 max_len = max(max_len, len(str(cell.value)))
    #         except:
    #             pass
    #     ws_summary.column_dimensions[col_letter].width = max_len + 2

    # # จัดรูปแบบ statistics sheet
    # if "statistics" in wb.sheetnames:
    #     ws_stats = wb["statistics"]
    #     for col in range(1, ws_stats.max_column + 1):
    #         col_letter = get_column_letter(col)
    #         ws_stats.column_dimensions[col_letter].width = 20

    wb.save(output_path)
    
    # รายงานสถิติ
    total_points_coords = sum(info['count_by_coords'] for info in redline_summary.values())
    total_points_details = sum(info['count_by_details'] for info in redline_summary.values())
    total_matches = sum(info['total_matches'] for info in redline_summary.values())
    
    logging.info("บันทึกผลเป็น Excel ที่: %s", output_path)
    logging.info("สถิติ: จุดที่ไม่ซ้ำ (coords)=%d, จุดที่ไม่ซ้ำ (details)=%d, matches รวม=%d", 
                total_points_coords, total_points_details, total_matches)
    
    return output_path